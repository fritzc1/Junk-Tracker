"""
Excel Box Contents Splitter

Reads an Excel spreadsheet where one column contains newline-separated item lists
(box contents) and splits each item into its own row, duplicating the other columns.

Usage:
    python split_box_contents.py input_spreadsheet.xlsx

Optional:
    python split_box_contents.py input_spreadsheet.xlsx -o output.xlsx
"""

import argparse
import sys
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install it with:")
    print("  pip install openpyxl")
    sys.exit(1)


def load_workbook(file_path):
    """Load an Excel workbook and return the active worksheet."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        return wb, ws
    except Exception as e:
        print(f"Error loading file: {e}")
        sys.exit(1)


def get_headers(ws):
    """Extract headers from the first row."""
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value is not None else "")
    return headers


def print_column_selection(headers):
    """Display available columns for user selection."""
    print("\nAvailable columns:")
    print("-" * 40)
    for idx, header in enumerate(headers):
        display_name = header if header else f"(Column {idx + 1})"
        print(f"  [{idx}] {display_name}")
    print()


def get_user_column_choice(headers):
    """Prompt user to select the contents list column."""
    print_column_selection(headers)

    while True:
        try:
            choice = input("Enter the column number for 'Contents List': ").strip()
            # Support both 0-indexed and 1-indexed input
            if choice:
                col_idx = int(choice)
                # If user enters a 1-based index that matches len, adjust
                if col_idx == len(headers):
                    col_idx = col_idx - 1
                if 0 <= col_idx < len(headers):
                    return col_idx
            print(f"Please enter a number between 0 and {len(headers) - 1}.")
        except ValueError:
            print("Invalid input. Please enter a number.")


def parse_contents(contents_cell):
    """
    Parse the contents cell by splitting on newlines.
    Returns a list of non-empty item strings.
    """
    if contents_cell is None:
        return []

    content_str = str(contents_cell).strip()
    if not content_str:
        return []

    # Split on various newline patterns (\r\n, \n, \r)
    items = re.split(r'\r?\n|\r', content_str)
    # Filter out empty strings and whitespace-only entries
    items = [item.strip() for item in items if item.strip()]
    return items


def split_rows(ws, contents_col_idx):
    """
    Process the worksheet and split rows based on the contents column.

    Returns:
        tuple: (headers, new_rows) where new_rows is a list of lists
    """
    headers = get_headers(ws)
    new_rows = []
    total_items = 0
    rows_processed = 0
    empty_content_rows = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Convert row to list
        row_data = list(row)

        contents_cell = row_data[contents_col_idx]
        items = parse_contents(contents_cell)

        rows_processed += 1

        if not items:
            # If no items found, keep the original row but blank out contents
            empty_content_rows += 1
            cleaned_row = list(row_data)
            cleaned_row[contents_col_idx] = ""
            new_rows.append(cleaned_row)
            continue

        total_items += len(items)

        for item in items:
            new_row = list(row_data)
            new_row[contents_col_idx] = item
            new_rows.append(new_row)

    print(f"\nProcessing Summary:")
    print(f"  Rows processed: {rows_processed}")
    output_columns = list(enumerate(headers))
    print(f"  Items extracted: {total_items}")
    print(f"  Empty content rows (kept blank): {empty_content_rows}")
    print(f"  Total output rows: {len(new_rows)}")

    return headers, new_rows


def save_output(headers, new_rows, output_path):
    """Save the processed data to a new Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data rows
    for row_idx, row_data in enumerate(new_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    # Handle multi-line content
                    if '\n' in str(cell.value):
                        cell_len = max(len(line) for line in str(cell.value).split('\n'))
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
        ws.column_dimensions[column_letter].width = adjusted_width

    try:
        wb.save(str(output_path))
        print(f"\nOutput saved to: {output_path}")
    except Exception as e:
        print(f"Error saving file: {e}")
        sys.exit(1)


def generate_default_output_path(input_path):
    """Generate a default output filename."""
    path = Path(input_path)
    stem = path.stem
    suffix = path.suffix
    output_name = f"{stem}_split{suffix}"
    return path.parent / output_name


def main():
    parser = argparse.ArgumentParser(
        description="Split Excel rows by newline-separated contents in a chosen column."
    )
    parser.add_argument("input_file", help="Path to the input Excel file (.xlsx)")
    parser.add_argument(
        "-o", "--output",
        help="Path for the output Excel file (default: <input>_split.xlsx)"
    )

    args = parser.parse_args()

    input_path = Path(args.input_file)
    if not input_path.exists():
        print(f"Error: File not found: {input_path}")
        sys.exit(1)

    if input_path.suffix.lower() not in ['.xlsx']:
        print("Note: This tool works best with .xlsx files.")
        print("For .xls files, consider converting to .xlsx first.")

    print(f"Loading: {input_path}")
    wb, ws = load_workbook(input_path)

    if ws.max_row < 2:
        print("Error: Worksheet appears to be empty or has no data rows.")
        sys.exit(1)

    headers = get_headers(ws)
    print(f"Found {ws.max_row - 1} data rows with {len(headers)} columns.")

    # Let user select the contents column
    contents_col_idx = get_user_column_choice(headers)
    selected_header = headers[contents_col_idx] if headers[contents_col_idx] else f"Column {contents_col_idx + 1}"
    print(f"\nSelected: '{selected_header}' as the Contents List column.")

    # Process and split rows
    print("\nProcessing...")
    headers, new_rows = split_rows(ws, contents_col_idx)

    # Determine output path
    output_path = Path(args.output) if args.output else generate_default_output_path(input_path)

    # Confirm before saving
    print(f"\nOutput file: {output_path}")
    confirm = input("Save results? (y/n): ").strip().lower()
    if confirm in ['y', 'yes', '']:
        save_output(headers, new_rows, output_path)
        print("Done!")
    else:
        print("Cancelled.")


if __name__ == "__main__":
    main()
